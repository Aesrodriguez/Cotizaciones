"""
Parser de archivos Excel de detalle de pagos/transferencias Bancolombia.
Soporta dos hojas:
  - Logs_pagos         → pagos a proveedores (PROV) y nómina (NOMI)
  - Logs_transferencias → transferencias cuenta inscrita (TRCI) y no inscrita (TRCN)
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl


# ── Helpers ───────────────────────────────────────────────────────────────────

def _s(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ('none', 'nan', 'n/a', '-') else None


def _d(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _dec(val: Any) -> Decimal | None:
    if val is None:
        return None
    try:
        s = str(val).replace(',', '').replace('$', '').replace(' ', '').strip()
        return Decimal(s) if s else None
    except InvalidOperation:
        return None


def _proceso(val: Any) -> str | None:
    s = _s(val)
    if not s:
        return None
    # Remove decimals if Excel stored as float (e.g. 63518073.0)
    try:
        return str(int(float(s)))
    except (ValueError, OverflowError):
        return s


# ── Parsers por hoja ──────────────────────────────────────────────────────────

def _parse_pagos(ws) -> list[dict]:
    """Hoja Logs_pagos — 27 columnas A:AA"""
    rows = []
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row_vals: list, name: str) -> Any:
        try:
            idx = next(i for i, h in enumerate(headers) if h and name.lower() in str(h).lower())
            return row_vals[idx]
        except StopIteration:
            return None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        proceso = _proceso(col(row, 'Proceso'))
        if not proceso:
            continue
        rows.append({
            'proceso':                     proceso,
            'servicio':                    _s(col(row, 'Servicio')),
            'nombre_servicio':             _s(col(row, 'Nombre Servicio')),
            'descripcion_pago':            _s(col(row, 'Descripcion Pago')),
            'tipo_producto_origen':        _s(col(row, 'Tipo Producto Origen')),
            'producto_origen':             _s(col(row, 'Producto Origen')),
            'fecha_pago_actualizacion':    _d(col(row, 'Fecha Pago y Actualizacion')),
            'estado':                      _s(col(row, 'Estado')),
            'fecha_creacion':              _d(col(row, 'Fecha Creacion')),
            'usuario_creacion':            _s(col(row, 'Usuario Creacion')),
            'usuario_aprueba_1':           _s(col(row, 'Usuario Aprueba 1')),
            'usuario_aprueba_2':           _s(col(row, 'Usuario Aprueba 2')),
            'nit_destino':                 _s(col(row, 'NIT Destino')),
            'nombre_destinatario':         _s(col(row, 'Nombre Destinatario')),
            'tipo_producto_destino':       _s(col(row, 'Tipo Producto Destino')),
            'producto_destino':            _s(col(row, 'Producto Destino')),
            'numero_convenio':             _s(col(row, 'Numero Convenio')),
            'fecha_pago':                  _d(col(row, 'Fecha Pago')),
            'referencia_destino':          _s(col(row, 'Referencia Destino')),
            'numero_referencia_destino':   _s(col(row, 'Numero Referencia Destino')),
            'monto':                       _dec(col(row, 'Monto Total Destino')),
            'banco_destino':               _s(col(row, 'Banco Destino')),
            'estado_registro':             _s(col(row, 'Estado Registro')),
            'causal_rechazo':              _s(col(row, 'Causal Rechazo')),
        })
    return rows


def _parse_transferencias(ws) -> list[dict]:
    """Hoja Logs_transferencias — 20 columnas A:T"""
    rows = []
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row_vals: list, name: str) -> Any:
        try:
            idx = next(i for i, h in enumerate(headers) if h and name.lower() in str(h).lower())
            return row_vals[idx]
        except StopIteration:
            return None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        proceso = _proceso(col(row, 'Proceso'))
        if not proceso:
            continue
        rows.append({
            'proceso':                  proceso,
            'servicio':                 _s(col(row, 'Servicio')),
            'nombre_servicio':          _s(col(row, 'Nombre Servicio')),
            'tipo_producto_origen':     _s(col(row, 'Tipo Producto Origen')),
            'producto_origen':          _s(col(row, 'Producto Origen')),
            'fecha_pago_actualizacion': _d(col(row, 'Fecha Pago y Actualizacion')),
            'nit_destino':              _s(col(row, 'NIT Destino')),
            'nombre_destino':           _s(col(row, 'Nombre Destino')),
            'fecha_creacion':           _d(col(row, 'Fecha de creacion')),
            'usuario_creacion':         _s(col(row, 'Usuario Creacion')),
            'usuario_aprueba':          _s(col(row, 'Usuario Aprueba')),
            'fecha_modificacion':       _d(col(row, 'Fecha Modificacion')),
            'tipo_producto_destino':    _s(col(row, 'Tipo producto Destino')),
            'producto_destino':         _s(col(row, 'Producto Destino')),
            'banco_destino':            _s(col(row, 'Banco Destino')),
            'monto':                    _dec(col(row, 'Monto Destino')),
            'estado':                   _s(col(row, 'Estado')),
            'causal_rechazo':           _s(col(row, 'Causal Rechazo')),
        })
    return rows


# ── Función principal ─────────────────────────────────────────────────────────

def parse_detalle_pago_xlsx(content: bytes) -> dict:
    """
    Lee un Excel de detalle de pagos/transferencias Bancolombia.
    Devuelve:
      { pagos: [...], transferencias: [...], hojas_encontradas: [...] }
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = [s.lower() for s in wb.sheetnames]

    pagos = []
    transferencias = []
    hojas = []

    for name in wb.sheetnames:
        nl = name.lower()
        if 'pago' in nl and 'transfer' not in nl:
            pagos = _parse_pagos(wb[name])
            hojas.append(name)
        elif 'transfer' in nl:
            transferencias = _parse_transferencias(wb[name])
            hojas.append(name)

    if not pagos and not transferencias:
        raise ValueError(
            f"No se encontraron hojas reconocibles. "
            f"Se esperaba 'Logs_pagos' y/o 'Logs_transferencias'. "
            f"Hojas en el archivo: {', '.join(wb.sheetnames)}"
        )

    return {
        'pagos':           pagos,
        'transferencias':  transferencias,
        'hojas_encontradas': hojas,
    }
